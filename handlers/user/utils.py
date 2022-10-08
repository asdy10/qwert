import base64
import time
from datetime import datetime

import pandas as pd

from utils.db_get_info.get_set_info_db import get_qr, get_seller_info, create_seller_info
from PIL import Image
from io import BytesIO
import requests
import openpyxl
from openpyxl import Workbook
import urllib.request

from utils.wb_api.work_wb_api import get_image_url_product, get_ip_ooo


def decode_img(qr):
    name = f'res{str(time.time()).replace(".", "")}.png'
    with open(name, 'wb') as f:
        f.write(base64.b64decode(qr))
    return name


def create_table_pvz(buyouts):
    res = {}
    for i in buyouts:
        g = [j for j in i]
        try:
            res[i[5]] = res[i[5]] + [g]
        except:
            res[i[5]] = [g]
    result = []
    for i in res:
        for j in res[i]:
            try:
                qr = get_qr(j[9])
                g = [i for i in j]
                g.append(qr)
                result.append(g)
            except:
                pass
        result.append([' '] * 12)

    df = pd.DataFrame({'idx': [i[1] for i in result],
                       'link': [i[2] for i in result],
                       'keyword': [i[3] for i in result],
                       'count': [i[4] for i in result],
                       'address': [i[5] for i in result],
                       'date': [i[6] for i in result],
                       'status': [i[7] for i in result],
                       'review': [i[8] for i in result],
                       'bid': [i[9] for i in result],
                       'price': [i[10] for i in result]
                       })
    #print(8)
    table_name = f'tables/ready_for_pickup_{round(time.time())}.xlsx'
    df.to_excel(table_name, sheet_name='Result', index=False)
    return table_name


def create_table_default(res):
    receipt_arr = []
    for i in res:
        if i[11] not in [0, None, '0']:
            print(3, i[11])
            receipt_arr.append([i[11].split(";")[0], i[11].split(";")[1]])
        else:
            receipt_arr.append([0, 0])
    df = pd.DataFrame({'idx': [i[1] for i in res],
                       'link': [i[2] for i in res],
                       'keyword': [i[3] for i in res],
                       'count': [i[4] for i in res],
                       'address': [i[5] for i in res],
                       'date': [i[6] for i in res],
                       'status': [i[7] for i in res],
                       'review': [i[8] for i in res],
                       'bid': [i[9] for i in res],
                       'price': [i[10] for i in res],
                       'rid': [i[0] for i in receipt_arr],
                       'receipt': [i[1] for i in receipt_arr]
                       })
    table_name = f'tables/all_{round(time.time())}.xlsx'
    df.to_excel(table_name, sheet_name='Result', index=False)
    return table_name


def create_dict(buyouts):
    res = {}
    for i in buyouts:
        g = [j for j in i]
        try:
            res[i[5]] = res[i[5]] + [g]
        except:
            res[i[5]] = [g]
    return res


def get_img(filename, size=(100, 100)):
    img = Image.open(filename)
    if size:
        img = img.resize(size)
    temp = BytesIO()
    img.save(temp, format="png")
    temp.seek(0)
    return Image.open(temp)


def save_image(url):
    #print(2, url)
    name = url.split('/')[-4]
    img = urllib.request.urlopen(url).read()
    out = open(f"product_images\\{name}.png", "wb")
    out.write(img)


def insert_row(ws, buyout, size=(200, 200)):
    print(1)
    pid = buyout[2].split('/')[4]
    img = ''
    try:
        img = openpyxl.drawing.image.Image(get_img(f"product_images\\{pid}.png", size))
        print(2)
    except:#no image
        try:
            url = get_image_url_product(pid)
            print('pid', pid)
            save_image(url)
            time.sleep(0.5)
            img = openpyxl.drawing.image.Image(get_img(f"product_images\\{pid}.png", size))
        except Exception as e:
            print(e)
    if buyout[11] not in [0, None, '0']:
        #print(6, buyout[11])
        print(3)
        rid = buyout[11].split(";")[0]
        receipt = buyout[11].split(";")[1]
    else:
        rid, receipt = 0, 0
    print(4)
    seller = get_seller_info(pid)
    if not seller:
        print(5)
        r = get_ip_ooo(pid)
        print(6)
        r['pid'] = pid
        create_seller_info(r)
        seller = get_seller_info(pid)
        print(7)
    pid, seller_name, ogrn, inn = seller
    row_num = ws.max_row + 1
    ws[f"B{row_num}"] = buyout[2] # link
    ws[f"C{row_num}"] = buyout[2].split('/')[4]
    ws[f"D{row_num}"] = seller_name
    ws[f"E{row_num}"] = buyout[6].split(' ')[0]
    ws[f"F{row_num}"] = buyout[6].split(' ')[1]
    ws[f"G{row_num}"] = buyout[10]  # price
    ws[f"H{row_num}"] = buyout[4]  # count
    ws[f"I{row_num}"] = buyout[7]  # status
    ws[f"J{row_num}"] = rid
    ws[f"K{row_num}"] = receipt
    ws[f"L{row_num}"] = buyout[5]  # address
    ws[f"M{row_num}"] = buyout[1]  # idx
    ws[f"N{row_num}"] = buyout[8]  # review
    ws[f"O{row_num}"] = buyout[9]  # bid
    cell_addr = f"A{row_num}"
    if img:
        img.anchor = cell_addr
        ws.add_image(img)

    ws.row_dimensions[row_num].height = int(size[1] * .8)
    #ws.column_dimensions["A"].width = int(size[0] * .15)
    return ws


def set_name_colon(ws):

    ws[f"A1"] = 'Фото'
    ws.column_dimensions["A"].width = 8
    ws[f"B1"] = 'Ссылка'
    ws.column_dimensions["B"].width = 45
    ws[f"C1"] = 'Артикул'
    ws.column_dimensions["C"].width = 13
    ws[f"D1"] = 'Продавец'
    ws.column_dimensions["D"].width = 16
    ws[f"E1"] = 'Дата выкупа'
    ws.column_dimensions["E"].width = 13
    ws[f"F1"] = 'Время выкупа'
    ws.column_dimensions["F"].width = 13
    ws[f"G1"] = 'Цена'
    ws.column_dimensions["G"].width = 6
    ws[f"H1"] = 'Количество'
    ws.column_dimensions["H"].width = 5
    ws[f"I1"] = 'Статус'
    ws.column_dimensions["I"].width = 19
    ws[f"J1"] = 'Номер заказа'
    ws.column_dimensions["J"].width = 13
    ws[f"K1"] = 'Чек'
    ws.column_dimensions["K"].width = 60
    ws[f"L1"] = 'ПВЗ'
    ws.column_dimensions["L"].width = 60
    ws[f"M1"] = 'idx'
    ws.column_dimensions["M"].width = 14
    ws[f"N1"] = 'Был ли отзыв'
    ws.column_dimensions["N"].width = 13
    ws[f"O1"] = 'bid'
    ws.column_dimensions["O"].width = 5

    return ws


def sort_by_date(b):
    date_time = datetime.strptime(b[6], "%d.%m.%Y %H:%M:%S")
    return date_time.timestamp()


def create_table_with_images(buyouts):
    buyouts.sort(key=sort_by_date)
    size = (50, 50)
    wb = Workbook()
    ws = wb.active
    ws = set_name_colon(ws)
    url_list = []
    for b in buyouts:
        ws = insert_row(ws, b, size=size)
    table_name = f'tables/all_{round(time.time())}.xlsx'
    print('save')
    wb.save(table_name)
    return table_name
