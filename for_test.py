import asyncio
import base64
import json
import os
import pickle
import time
from datetime import datetime
from threading import Thread

import pandas as pd
import requests
from bs4 import BeautifulSoup
from selenium.webdriver import ActionChains
# from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By

from browser.cart_products import set_payment_method, get_price, check_is_product_in_cart
from browser.cookies_browser import get_open_browser
from browser.liked_products import delete_from_liked
from browser.registration_browser import set_name_account
from loader import db
#from utils.android.android_connector import device
from utils.db_get_info.get_set_info_db import *
from utils.proxy_configurator.proxy_configurator import change_proxy_ip
from utils.wb_api.work_wb_api import get_data_delivery_receipts, get_image_url_product, get_delivery_req, \
    get_normal_time, get_ip_ooo
from utils.yamaps.ya_maps import get_address_and_photo_not_async


def get_cookies(phone, driver):
    driver.get('https://www.wildberries.ru/security/login?returnUrl=https%3A%2F%2Fwww.wildberries.ru%2F')
    s = input()
    time.sleep(1)
    pickle.dump(driver.get_cookies(), open(f'cookies_{phone}', 'wb'))
    time.sleep(5)
    driver.close()


def set_cookies(driver, phone):
    driver.get('https://www.wildberries.ru/')
    time.sleep(3)
    dir = os.path.abspath(os.curdir)
    for cookie in pickle.load(open(f'browser\\cookies\\cookies_{phone}', 'rb')):
        driver.add_cookie(cookie)
    time.sleep(3)
    driver.refresh()
    time.sleep(10)
    return driver


def check_payment_browser(driver, idx=0):
    try:
        el = driver.find_element(By.CLASS_NAME, 'popup__header').text
        print(el)
        check = 0
        if el:
            while el == 'QR-код для оплаты' and check < 600:
                try:
                    el = driver.find_element(By.CLASS_NAME, 'popup__header').text
                except:
                    time.sleep(10)
                    check += 10
                    try:

                        el2 = driver.find_element(By.CLASS_NAME, 'order-fail').text
                        print(el2)
                        el = driver.find_element(By.CLASS_NAME, 'order-fail__text').text
                    except:
                        el = 'Good sale'
                time.sleep(10)
                check += 10
            if check < 600:
                page = driver.page_source
                soup = BeautifulSoup(page, "html.parser")
                # with open('test.txt', 'w', encoding='utf-8') as f:
                #     f.write(page)
                tit = ''
                for tag in soup.findAll('div', class_='order-fail'):
                    tit = tag.find('h1', class_='order-fail__title').text
                if tit == 'К сожалению, Ваш платёж не удался!':
                    payment = 'Платеж не удался, заказ отменен'
                else:
                    payment = 'Оплата успешна'
                #loop = asyncio.get_event_loop()
                #loop.run_until_complete(send_fail_payment(idx.split('_')[0], payment))
                return driver, payment
            else:
                return driver, 'error check payment'
        else:
            return driver, 'error no name'
    except:
        return driver, 'error check payment'


from selenium import webdriver
from seleniumwire import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import logging

from yoomoney import Authorize
from yoomoney import Client
from yoomoney import Quickpay


def create_link_for_payment(receiver, label, summa):
    quick_pay = Quickpay(
        receiver=receiver,
        quickpay_form="shop",
        targets="Sponsor this project",
        paymentType="SB",
        sum=summa,
        label=label
    )
    # print(quick_pay.redirected_url)
    return quick_pay.base_url


def check_payment(label, token):
    client = Client(token)
    history = client.operation_history()
    print(history.operations)
    for i in history.operations:
        print(i.amount, i.status, i.label)
    if history.operations[0].status == 'success':
        return True
    else:
        return False


def get_token(client_id, redirect_uri):
    Authorize(
        client_id=client_id,
        redirect_uri=redirect_uri,
        scope=["account-info",
               "operation-history",
               "operation-details",
               "incoming-transfers",
               "payment-p2p",
               "payment-shop",
               ]
    )


def get_client_info(token):
    client = Client(token)
    user = client.account_info()
    print("Account number:", user.account)
    print("Account balance:", user.balance)
    print("Account currency code in ISO 4217 format:", user.currency)
    print("Account status:", user.account_status)
    print("Account type:", user.account_type)
    print("Extended balance information:")
    for pair in vars(user.balance_details):
        print("\t-->", pair, ":", vars(user.balance_details).get(pair))
    print("Information about linked bank cards:")
    cards = user.cards_linked
    if len(cards) != 0:
        for card in cards:
            print(card.pan_fragment, " - ", card.type)
    else:
        print("No card is linked to the account")


import warnings

warnings.filterwarnings("ignore")


def clear_cart(driver):
    actions = ActionChains(driver)
    k = driver.find_elements_by_class_name('accordion__list-item')
    for i in k:
        actions.move_to_element(i).perform()
        time.sleep(1)
        clear = driver.find_element_by_class_name('btn__del')
        actions.move_to_element(clear).perform()
        time.sleep(1)
        clear.click()
        time.sleep(2)
    return driver


import cv2
import numpy as np


def qrdecode(image):
    inputImage = cv2.imread(image)
    qrDecoder = cv2.QRCodeDetector()
    data, _, rectifiedImage = qrDecoder.detectAndDecode(inputImage)
    rectifiedImage = np.uint8(rectifiedImage)

    cv2.imwrite("Rectified QRCode.png", rectifiedImage)
    # time.sleep(10)
    return data


def asdfg():
    # qr = qrdecode('qrfull.png')
    # print(qr)
    phone = '79876146123'
    proxy = ''
    user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.0.0 Safari/537.36'
    driver = get_open_browser(proxy, user_agent)
    driver.get('https://google.com')
    name = datetime.today().strftime("%H%M%S")
    driver.get_screenshot_as_file(f'{name}.png')
    driver.close()
    # get_cookies(phone, driver)
    # set_cookies(driver, phone)
    # clear_cart(driver)
    s = 'Итого\n286 ₽'
    # price = s[5:-2].replace(' ', '')
    # print(price)
    label = ''
    token = '4100117882116084.2A2303F16017D011565A3DDD31C43BD5880470A137F0C28F1350BEB93D06FD00EECC4AA91975D9DD4AD21B368F895623F355A8A47E9A0C3D8C2DFC57EDEB4B9996AED9B503A2F813CC2AAB27D7C42CB01F62DDD9A6603185974460E126DD32EE3DADF023F4715FB226F68F013C96720EC345F49AF8429D2398422A84DABF160B'
    # check_payment(label, token)
    client_id = '3C8475003D576AD1BFFA9AB9DBBEFD2BA5201E32512AA4FB872E31D141C72B95'
    receiver = "410011773918938"
    label = "a1b2c3d4e5"
    token = '410011773918938.DA56E4499B63A71A2760B16871BF3F11B1E9B248119A512E14E7B84BFE8DB58559FCD875943C9D4AD065B3FEAB5FAEDC62C199050085BF0DCAE4F4DB0C90151CD2F3B22190BB2544DF1667BA8EA560DB51399C10992B665E7A5E6CB480F4D37B47B14D713DCD577CE57D8BA16167C70182CD34530195397D5CA99B4626DC8CFF'
    # check_payment(label, token)
    redirect_uri = 'https://t.me/original_purple_cow_bot'
    # get_token(client_id, redirect_uri)
    user_id = 383552200


def foo():
    time.sleep(1)
    return 'foo23'


def test1():
    url = 'https://www.wildberries.ru/webapi/spa/modules/pickups'
    cookies = """cookie: BasketUID=62cf91b6-c80d-4fee-ac6e-e44c89c434d6; _gcl_au=1.1.2035184656.1654951773; _wbauid=3717703611654951772; _ga=GA1.2.937760077.1654951773; route=1654951774.007.6749.762287|bc5b49a034e8c65a5a36b6c2a453bfa0; __wba_s=1; ___wbu=b511796f-c526-4617-89c6-769b1bf009a0.1654951780; ___wbu=5ea8e470-400c-4aa4-bb76-6d3b534cc520.1656610871; __bsa=basket-ru-42; WILDAUTHNEW_V3=DD00407934B970F1B869FD8F066DA9D299D73F851E093E83ED729300402B70D2904FAA4F10D00E4D0941BC4BF3C32C3ECCDDA58E7531C7BD1D21803DA748EB4F371FFE391C08A611A409906AEA989AA35C22E6164E4B44E33C474C721F4C88F692C0A1522B8D3657143D08B7458285A579C9ECC5C16941466DAE141B4CD4CC65AB8911A1C6CD7F728A0F516654739F4CF2FEF3AAA6D93D2B086C5483DD668AAC28C5A68412F417F7616956C3B3F284B23DC5DF73EEEB1627F7E3143760F6D3E2758BF98CA5C40ADA44A0AA1FF74CC855B737FC77FBFF13A8F654AA2B05CD817D37C68333B8E225E73A814B5F9C7557C3AA8B9C6DE1A75FD93207C34050D87DD93A654528720E7D2F57530CAA5F5AB022C1D44204E4AA7D8D27CD324BC1EC74025796D50358BDD7285B50B63122C55448E5C350851EFE2C2BC48BBBE584FF1B3F39364F2BD97FADFABCDEC15B40F52757F4E4091B; _wbSes=CfDJ8GWigffatjpAmgU4Ds4%2BnhvTeRrv9OOB29bMQ%2B26lJnXf0WZP291XMBQbctY%2F%2B%2F0NaIf2mbbr5nvvOlzROJU%2B%2B1SaJnmJrOdfouL2mRqUmGLEm7J%2BGzYgyvxc2b39xGcpCxsvvnWRL%2BOQ2AyYSeyJglbCKqBRzY8wydDwdoNW7VF; __bsa=basket-ru-42; __wbl=cityId%3D0%26regionId%3D0%26city%3D%D1%81%20%D0%9D%D0%B0%D0%B3%D0%B0%D0%B5%D0%B2%D0%BE%2C%20%D0%9D%D0%BE%D0%B2%D0%B0%D1%8F%20%D0%A3%D0%BB%D0%B8%D1%86%D0%B0%201%26phone%3D84957755505%26latitude%3D54%2C625223%26longitude%3D56%2C104452%26src%3D1; __store=117673_122258_122259_117986_1733_117501_507_3158_120762_204939_159402_2737_130744_686_1193_124731_121709; __region=64_83_4_38_80_33_70_82_86_30_69_22_66_31_40_1_48; __pricemargin=1.0--; __cpns=2_12_7_3_6_18_21; __sppfix=; __dst=-1075831_-77677_-398551_12358502; ncache=117673_122258_122259_117986_1733_117501_507_3158_120762_204939_159402_2737_130744_686_1193_124731_121709%3B64_83_4_38_80_33_70_82_86_30_69_22_66_31_40_1_48%3B1.0--%3B2_12_7_3_6_18_21%3B%3B-1075831_-77677_-398551_12358502; um=uid%3Dw7TDssOkw7PCu8K5wrfCsMK5wrTCsMKzwrA%253d%3Aproc%3D100%3Aehash%3D8249bb93d039be873a02732ef40a8424; _gid=GA1.2.1944336370.1660751634; __tm=1660822478; _dc_gtm_UA-2093267-1=1; ___wbs=089e40a2-dd96-47b7-935f-8fc00a59a851.1660811679"""
    headers = """accept: */*
accept-encoding: gzip, deflate, br
accept-language: ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7
referer: https://www.wildberries.ru/services/besplatnaya-dostavka?desktop=1
sec-ch-ua: "Chromium";v="104", " Not A;Brand";v="99", "Google Chrome";v="104"
sec-ch-ua-mobile: ?0
sec-ch-ua-platform: "Windows"
sec-fetch-dest: empty
sec-fetch-mode: cors
sec-fetch-site: same-origin
user-agent: Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.0.0 Safari/537.36
x-requested-with: XMLHttpRequest
x-spa-version: 9.3.22"""
    arr_headers = {i.split(': ')[0]: i.split(': ')[1] for i in headers.split('\n')}
    arr_cook = {i.split('=')[0]: i.split('=')[1] for i in cookies.split('; ')}
    res = requests.get(url=url, headers=arr_headers, cookies=arr_cook)
    res = res.json()['value']['pickups']
    # print(res)
    coor_arr = []
    # for i in res:
    #     print(i)
    for i in res:
        x, y = i['coordinates']
        coor_arr.append([round(float(x), 2), round(float(y), 2)])

    print('wtf')
    return coor_arr


from onlinesimru import GetFree, GetRent, GetProxy, GetUser, GetNumbers, Driver


def main():
    token = 'Z91Ch3YbLF78cMs-WL4ehTn7-dBhM41X3-UG9aBbrN-pMg1Ky2v8Xmkc91'
    # client = GetUser()
    driver = Driver(token).numbers()
    # tariffs = driver.tariffs()
    # tariffs = tariffs['7']['services']
    # #{'count': 1302, 'popular': False, 'code': 7, 'price': 6, 'id': 18, 'service': 'Wildberries', 'slug': 'wildberries'}
    # for t in tariffs:
    #     if 'wild' in t:
    #         print(tariffs[t])
    tzid = driver.get(service='Wildberries', number=True)
    print(tzid, tzid['number'])
    code = driver.wait_code(tzid['tzid'])
    print(code)
    # print(tariffs)


def rename():
    with open('data/lastname_woman2.txt', 'r', encoding='utf-8') as f:
        arr = f.read().split('\n')
    new_arr = []
    for i in arr:
        new_arr.append(i.split(' ')[2])
    with open('data/lastname_woman.txt', 'w', encoding='utf-8') as f:
        f.write('\n'.join(new_arr))
    with open('data/lastname_man2.txt', 'r', encoding='utf-8') as f:
        arr = f.read().split('\n')
    new_arr = []
    for i in arr:
        new_arr.append(i.split(' ')[2])
    with open('data/lastname_man.txt', 'w', encoding='utf-8') as f:
        f.write('\n'.join(new_arr))


def get_qr(driver):
    el = driver.find_elements(By.CLASS_NAME, 'delivery-qr__code')
    for i in el:
        try:
            res = i.get_attribute('src')
            return res
        except:
            pass


# def get_check_id(link):
#     res = requests.get(link)
#     #print(res.text)
#     so0up = BeautifulSoup(res.text, "html.parser")
#     ids = [tag['id'] for tag in soup.findAll('div', class_='product-card')]



from PIL import Image
from io import BytesIO
import requests
import openpyxl
from openpyxl import Workbook


def get_img(filename, size=(100, 100)):
    # r = requests.get(url, stream=True)
    # if not r.ok:
    #     r.raise_for_error()
    # r.raw.decode_content = True
    # img = Image.open(r.raw)
    # if size:
    #     img = img.resize(size)
    #return Image.open(temp)
    # img = Image.open(filename)
    # # if size:
    # #     img = img.resize(size)
    # # temp = BytesIO()
    # # img.save(temp, format="png")
    # # temp.seek(0)
    # #img.show()
    # return img
    img = Image.open(filename)
    if size:
        img = img.resize(size)
    temp = BytesIO()
    img.save(temp, format="png")
    temp.seek(0)
    return Image.open(temp)

import urllib.request


def save_image(url):
    name = url.split('/')[-4]
    img = urllib.request.urlopen(url).read()
    out = open(f"product_images\\{name}.png", "wb")
    out.write(img)


def insert_row(ws, buyout, size=(200, 200)):
    name = buyout[2].split('/')[-2]

    img = openpyxl.drawing.image.Image(get_img(f"product_images\\{name}.png", size))
    if buyout[11] not in [0, None, '0']:
        rid = buyout[11].split(";")[0]
        receipt = buyout[11].split(";")[1]
    else:
        rid, receipt = 0, 0
    row_num = ws.max_row + 1
    ws[f"B{row_num}"] = buyout[2]  # link
    ws[f"C{row_num}"] = buyout[6]  # date
    ws[f"D{row_num}"] = buyout[10]  # price
    ws[f"E{row_num}"] = buyout[4]  # count
    ws[f"F{row_num}"] = buyout[3]  # keywords
    ws[f"G{row_num}"] = buyout[7]  # status
    ws[f"H{row_num}"] = rid
    ws[f"I{row_num}"] = receipt
    ws[f"J{row_num}"] = buyout[5]  # address
    ws[f"K{row_num}"] = buyout[1] #idx
    ws[f"L{row_num}"] = buyout[8] #review
    ws[f"M{row_num}"] = buyout[9] #bid

    cell_addr = f"A{row_num}"
    img.anchor = cell_addr
    ws.add_image(img)

    ws.row_dimensions[row_num].height = int(size[1] * .8)
    #ws.column_dimensions["A"].width = int(size[0] * .15)
    return ws


def set_name_colon(ws):

    ws[f"A1"] = 'image'
    ws.column_dimensions["A"].width = 8
    ws[f"B1"] = 'link'
    ws.column_dimensions["B"].width = 45
    ws[f"C1"] = 'date'
    ws.column_dimensions["C"].width = 19
    ws[f"D1"] = 'price'
    ws.column_dimensions["D"].width = 6
    ws[f"E1"] = 'count'
    ws.column_dimensions["E"].width = 5
    ws[f"F1"] = 'keyword'
    ws.column_dimensions["F"].width = 15
    ws[f"G1"] = 'status'
    ws.column_dimensions["G"].width = 19
    ws[f"H1"] = 'rid'
    ws.column_dimensions["H"].width = 13
    ws[f"I1"] = 'receipt'
    ws.column_dimensions["I"].width = 60
    ws[f"J1"] = 'address'
    ws.column_dimensions["J"].width = 60
    ws[f"K1"] = 'idx'
    ws.column_dimensions["K"].width = 14
    ws[f"L1"] = 'review'
    ws.column_dimensions["L"].width = 5
    ws[f"M1"] = 'bid'
    ws.column_dimensions["M"].width = 5
    return ws


if __name__ == '__main__':
    #get_delivery_req(9130879293)
    #res = get_delivery_req(9876146123)

    # res = get_all_buyouts_not_async()
    # for i in res:
    #     pid = i[2].split('/')[4]
    #     link = f'https://www.wildberries.ru/catalog/{pid}/detail.aspx?targetUrl=EX'
    #     idx = i[1]
    #     db.query('UPDATE buyouts SET link=? WHERE idx=?', (link, idx))
    # a = []
    # for i in res:
    #     if i[1] not in a:
    #         a.append(i[1])
    #     else:
    #         print(i[1])
    # print(res)
    # pid = '73257189'
    # r = get_ip_ooo(pid)
    # r['pid'] = pid
    # create_seller_info(r)

    # new_arr = []
    # arr = {}
    # try:
    #     buyouts = get_all_buyouts_not_async()
    #
    #     for b in buyouts:
    #         if b[9] not in new_arr:
    #             new_arr.append(b[9])
    #             arr[b[9]] = b
    #         else:
    #             #print(arr[b[9]])
    #             print(b[9], b)
    # except Exception as e:
    #     print(e)
    # size = (50, 50)
    # wb = Workbook()
    # ws = wb.active
    # ws = set_name_colon(ws)
    # url_list = []
    # for b in buyouts:
    #
    #     ws = insert_row(ws, b, size=size)
    # wb.save('test.xlsx')
    # df = pd.DataFrame({'idx': [i[1] for i in new_arr],
    #                    'image': [f'<img src="{i[12]}"/>' for i in new_arr],
    #
    #                    })
    # table_name = f'tables/image_{round(time.time())}.xlsx'
    # df.to_excel(table_name, sheet_name='Result', index=False)





    #create_product_image(pid, url)
    #print(url)
    user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.0.0 Safari/537.36'
    token = 'cfae5bc12c3d2460e778531cc1ce1069'
    proxy_key = 'a5b92dbc31115feb5fd40fdb09cb2e96'
    # link = 'https://qr.nspk.ru/BD100079OJJHGJAG83NRAERR9428F7H2?type=02&bank=100000000014&sum=97500&cur=RUB&crc=DB3B'
    # link2 = 'https://qr.nspk.ru/BD100079OJJHGJAG83NRAERR9428F7H2?type=02&bank=100000000014&sum=222200&cur=RUB&crc=DB3B'

    phone = '9876146123'
    proxy = ''
    phone = input('phone')

    change_proxy_ip(proxy_key, user_agent)
    proxy = 'UpEG3y:YP4Em8Af7RaS@mproxy.site:12954'


    driver = get_open_browser(proxy, phone)
    set_cookies(driver, phone)

    time.sleep(10)
    # res = check_is_product_in_cart(driver, '79324902')
    # print(res)
    # # driver.get('https://www.wildberries.ru/lk/favorites')
    # driver.get('https://business.kazanexpress.ru/')
    # time.sleep(10)
    # print(get_price(driver))

    # time.sleep(10)
    #driver = set_cookies(driver, '9876146123')
    #r = get_all_receipts()
    #print(r)
    phone = '9876146123'
    # print(phone[-4:])
    # print(type('asdf') == str)
    # r = get_delivery_req(phone, user_agent='', proxy=proxy)
    # print('r', r)
    # s = '2022-09-06T17:25:29Z'
    # date_, time_ = s.split('T')
    # y, m, d = date_.split('-')
    # time_ = time_[:-1]
    # result = get_normal_time(s)
    # print(result)
    #r = get_receipt_by_link('https://f.wb.ru/r/1610558796/b92422b8023823003f16b3a4208e258d')
    #print(r)
    #get_receipt_req('9876146123')
    #get_check_id('https://f.wb.ru/r/1610558796/b92422b8023823003f16b3a4208e258d')
    # driver = get_open_browser('d', 's')
    # driver = set_cookies(driver, '9876146123')
    # time.sleep(10)
    # get_qr(driver)
    # time.sleep(1000)
    # set_name_account(driver, 'male', 'Алексей', 'Ермолаев')
    # coor_arr = test1()
    # address = 'Химки, совхозная 10'
    # ad, photo, y, x = get_address_and_photo_not_async(address)
    # print(ad, round(float(x), 2), round(float(y), 2))
    # if [round(float(x), 2), round(float(y), 2)] in coor_arr:
    #     find = True
    #     print('true')
#     data = {
#   "qrcIds": [
#     "AD1000670LSS7DN18SJQDNP4B05KLJL2"
#   ]
# }
#     res = requests.put('https://sub.nspk.ru/api/payment/v1/qrc-status', json=data)
#     print(res)
#     print(res.text)
# if [float(x), float(y)] in coor_arr:
#     print('find wtf')
